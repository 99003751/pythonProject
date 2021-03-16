import os
import pandas as pd
import openpyxl as pxl
from xlsxwriter import Workbook
DataBase = os.path.isfile('Employee_data1.xlsx')
'''
xls = pd.ExcelFile('Employee_data1.xlsx')
df = pd.read_excel('Employee_data1.xlsx', 'Sheet1')
print(df)
'''

Name_list = []
Postal_list = []
Email_list = []
City_list = []
Address_list = []
Company_list = []
Country_list = []

searchName = input("Enter Name: ")

df = pd.read_excel('Employee_data1.xlsx', 'Sheet1')
#print(df)
for i in range(0, len(df)):
    if df.Name[i] == searchName:
        Name_list.append(df.Name[i])
        Email_list.append(df.Email_ID[i])
        Postal_list.append(df.Postal[i])


df = pd.read_excel('Employee_data1.xlsx', 'Sheet2')
#print(df)
for i in range(0, len(df)):
    if df.Name[i] == searchName:
        Address_list.append(df.Address[i])

df = pd.read_excel('Employee_data1.xlsx', 'Sheet3')
#print(df)
for i in range(0, len(df)):
    if df.Name[i] == searchName:
        City_list.append(df.City[i])

df = pd.read_excel('Employee_data1.xlsx', 'Sheet4')
#print(df)
for i in range(0, len(df)):
    if df.Name[i] == searchName:
        Company_list.append(df.Company[i])

df = pd.read_excel('Employee_data1.xlsx', 'Sheet5')
#print(df)
for i in range(0, len(df)):
    if df.Name[i] == searchName:
        Country_list.append(df.Country[i])
#print(data_list)


excel_book = pxl.load_workbook('Employee_data1.xlsx')
with pd.ExcelWriter('Employee_data1.xlsx', engine='openpyxl') as writer:
    writer.book = excel_book
    writer.sheets = {
        worksheet.title: worksheet
        for worksheet in excel_book.worksheets
    }
    data1 = pd.DataFrame({'Name': Name_list, 'Email ID': Email_list, 'Postal Address': Postal_list, 'Address': Address_list,
                          'City': City_list, 'Company': Company_list, 'Country': Country_list})
    data1.to_excel(writer, 'master', index=False)
    writer.save()


'''
from openpyxl import load_workbook
wb = load_workbook('Employee_data1.xlsx')

Data_List = ['']

if 'Sheet6' not in wb.sheetnames:
    ws = wb.create_sheet('Sheet6')
    #print("CREATING")
    s=ws.max_row 
    # variable to store max rows for sl num

    for i in range(1,6):
        ws.cell(row=s+1,column=i).value="heyyyy"
    wb.save('Employee_data1.xlsx')
'''
